"""
CSV parsing, daily aggregation, and Excel export for EV charging data.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KWH_PRICE_DEFAULT = 0.121

CLR_HEADER_BG = "1F4E79"
CLR_TITLE_BG  = "2E75B6"
CLR_ALT_ROW   = "D6E4F0"
CLR_TOTAL_BG  = "FFF2CC"
CLR_GRAND_BG  = "C6EFCE"
CLR_GRAND_FG  = "375623"
CLR_PRICE_BG  = "FFF2CC"

thin   = Side(style="thin",   color="AAAAAA")
medium = Side(style="medium", color="2E75B6")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_PRICE = Border(left=medium, right=medium, top=medium, bottom=medium)

PRICE_CELL = "Dashboard!$B$2"


def hdr_style(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def data_style(cell, bold=False, number_format=None, bg=None, fg="000000"):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = BORDER
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    if number_format:
        cell.number_format = number_format


def parse_csv(file_obj):
    """Parse a CSV file-like object; return a daily summary DataFrame."""
    df = pd.read_csv(file_obj)
    df.columns = df.columns.str.strip()

    time_col = ("Date (console local time)"
                if "Date (console local time)" in df.columns
                else "Date (UTC time)")

    df["_datetime"] = pd.to_datetime(df[time_col], errors="coerce")
    df["_date"]     = df["_datetime"].dt.date
    df["_month"]    = df["_datetime"].dt.to_period("M")
    df["Power Usage (kWh)"] = pd.to_numeric(df["Power Usage (kWh)"], errors="coerce").fillna(0)
    df["Charge Time (s)"]   = pd.to_numeric(df["Charge Time (s)"],   errors="coerce").fillna(0)
    df["Total Time (s)"]    = pd.to_numeric(df["Total Time (s)"],     errors="coerce").fillna(0)

    grp = df.groupby("_date").agg(
        sessions       = ("Power Usage (kWh)", "count"),
        total_kwh      = ("Power Usage (kWh)", "sum"),
        total_charge_s = ("Charge Time (s)",   "sum"),
        total_time_s   = ("Total Time (s)",     "sum"),
    ).reset_index()

    grp["month"]         = pd.to_datetime(grp["_date"]).dt.to_period("M")
    grp["month_str"]     = grp["month"].astype(str)
    grp["date_str"]      = grp["_date"].astype(str)
    grp["charge_hm"]     = grp["total_charge_s"].apply(lambda s: f"{int(s)//3600}h {(int(s)%3600)//60}m")
    grp["total_time_hm"] = grp["total_time_s"].apply(lambda s: f"{int(s)//3600}h {(int(s)%3600)//60}m")
    grp["day_name"]      = pd.to_datetime(grp["date_str"]).dt.strftime("%A")
    return grp


def build_dashboard_data(grp, kwh_price):
    """Return structured data for the web dashboard."""
    months = sorted(grp["month_str"].unique().tolist())
    month_data = {}

    DOW_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for m in months:
        mdf = grp[grp["month_str"] == m].copy().sort_values("date_str")
        mdf["cost"] = mdf["total_kwh"] * kwh_price

        days = []
        for _, row in mdf.iterrows():
            days.append({
                "date_str":      row["date_str"],
                "day_name":      row["day_name"],
                "sessions":      int(row["sessions"]),
                "total_kwh":     round(float(row["total_kwh"]), 3),
                "charge_hm":     row["charge_hm"],
                "total_time_hm": row["total_time_hm"],
                "cost":          round(float(row["cost"]), 4),
                "cum_cost":      0,  # kept for table but not charted
            })
        cum = 0
        for d in days:
            cum += d["cost"]
            d["cum_cost"] = round(cum, 4)

        # Day-of-week average cost (all data in file, not just this month)
        grp2 = grp.copy()
        grp2["cost"] = grp2["total_kwh"] * kwh_price
        dow_avg = grp2.groupby("day_name")["cost"].mean()
        dow_labels = [d for d in DOW_ORDER if d in dow_avg.index]
        dow_values = [round(float(dow_avg[d]), 4) for d in dow_labels]

        month_data[m] = {
            "days":           days,
            "total_kwh":      round(float(mdf["total_kwh"].sum()), 3),
            "total_cost":     round(float(mdf["total_kwh"].sum() * kwh_price), 4),
            "total_sessions": int(mdf["sessions"].sum()),
            "labels":         [d["date_str"] for d in days],
            "costs":          [d["cost"] for d in days],
            "kwh_vals":       [d["total_kwh"] for d in days],
            "dow_labels":     dow_labels,
            "dow_values":     dow_values,
        }

    grand_kwh  = round(float(grp["total_kwh"].sum()), 3)
    grand_cost = round(float(grand_kwh * kwh_price), 4)

    return {
        "months":         months,
        "month_data":     month_data,
        "grand_kwh":      grand_kwh,
        "grand_cost":     grand_cost,
        "grand_sessions": int(grp["sessions"].sum()),
        "kwh_price":      kwh_price,
    }


def build_excel(grp, kwh_price):
    """Build and return an Excel workbook as bytes."""
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    months = sorted(grp["month_str"].unique())

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "⚡  EV Charging Cost Report"
    c.font  = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=CLR_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Editable price cell
    ws["A2"].value = "⚡ kWh Price (editable):"
    ws["A2"].font  = Font(name="Arial", bold=True, size=11, color=CLR_HEADER_BG)
    ws["A2"].alignment = Alignment(vertical="center")
    pc = ws["B2"]
    pc.value         = kwh_price
    pc.number_format = "$#,##0.0000"
    pc.font          = Font(name="Arial", bold=True, size=12, color="375623")
    pc.fill          = PatternFill("solid", start_color=CLR_PRICE_BG)
    pc.alignment     = Alignment(horizontal="center", vertical="center")
    pc.border        = BORDER_PRICE
    ws.merge_cells("C2:G2")
    ws["C2"].value = "← Edit this cell to instantly recalculate ALL costs"
    ws["C2"].font  = Font(name="Arial", italic=True, size=9, color="7F7F7F")
    ws["C2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 28

    # Month dropdown
    ws["A3"].value = "📅  Select Month:"
    ws["A3"].font  = Font(name="Arial", bold=True, size=11, color=CLR_HEADER_BG)
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 26

    ws_list = wb.create_sheet("_MonthList")
    ws_list.sheet_state = "hidden"
    for i, ml in enumerate(months, 1):
        ws_list[f"A{i}"] = ml

    dv = DataValidation(type="list",
                        formula1=f"=_MonthList!$A$1:$A${len(months)}",
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    ws["B3"].value = months[0]
    ws["B3"].font  = Font(name="Arial", bold=True, size=11, color="1F4E79")
    ws["B3"].fill  = PatternFill("solid", start_color="EBF3FB")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B3"].border = BORDER_PRICE
    dv.add("B3")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44

    # Grand summary header
    ws.merge_cells("A5:G5")
    c = ws["A5"]
    c.value = "Grand Total Summary — All Months"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 22

    for col, h in enumerate(["Month", "Days Charged", "Total Sessions",
                              "Total kWh", "Total Cost ($)"], 1):
        hdr_style(ws.cell(row=6, column=col, value=h))

    data_row = 7
    grand_kwh_refs  = []
    grand_cost_refs = []

    for m in months:
        mdf = grp[grp["month_str"] == m]
        kwh = round(mdf["total_kwh"].sum(), 3)
        bg  = CLR_ALT_ROW if data_row % 2 == 0 else None
        kwh_ref  = f"D{data_row}"
        cost_ref = f"E{data_row}"
        grand_kwh_refs.append(kwh_ref)
        grand_cost_refs.append(cost_ref)

        for col, v in enumerate([m, len(mdf), int(mdf["sessions"].sum()), kwh], 1):
            data_style(ws.cell(row=data_row, column=col, value=v),
                       number_format="#,##0.000" if col == 4 else None, bg=bg)
        data_style(ws.cell(row=data_row, column=5,
                           value=f"={kwh_ref}*{PRICE_CELL}"),
                   number_format="$#,##0.0000", bg=bg)
        data_row += 1

    for col, v in enumerate(["GRAND TOTAL", "—", "—",
                              f"={'+'.join(grand_kwh_refs)}",
                              f"={'+'.join(grand_cost_refs)}"], 1):
        c = ws.cell(row=data_row, column=col, value=v)
        fmt = "$#,##0.0000" if col == 5 else ("#,##0.000" if col == 4 else None)
        data_style(c, bold=True, number_format=fmt, bg=CLR_GRAND_BG, fg=CLR_GRAND_FG)

    for i, w in enumerate([13, 13, 16, 12, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Per-month sheets
    for m in months:
        mdf  = grp[grp["month_str"] == m].copy().sort_values("_date")
        ws_m = wb.create_sheet(title=m)
        ws_m.sheet_view.showGridLines = False

        ws_m.merge_cells("A1:H1")
        c = ws_m["A1"]
        c.value = f"⚡  EV Charging Cost — {m}"
        c.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color=CLR_TITLE_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_m.row_dimensions[1].height = 32

        ws_m.merge_cells("A2:H2")
        ws_m["A2"].value = "kWh Price is set on the Dashboard (cell B2) — edit there to update all costs"
        ws_m["A2"].font  = Font(name="Arial", italic=True, size=9, color="7F7F7F")
        ws_m["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws_m.row_dimensions[2].height = 16

        for col, h in enumerate(["Date", "Day", "Sessions", "Total kWh",
                                  "Charge Time", "Total Time",
                                  "Cost per Day ($)"], 1):
            hdr_style(ws_m.cell(row=3, column=col, value=h))

        first = 4
        for r_idx, (_, row) in enumerate(mdf.iterrows(), first):
            bg = CLR_ALT_ROW if r_idx % 2 == 0 else None
            for col, v in enumerate([
                row["date_str"], row["day_name"], int(row["sessions"]),
                round(row["total_kwh"], 3), row["charge_hm"], row["total_time_hm"]
            ], 1):
                data_style(ws_m.cell(row=r_idx, column=col, value=v),
                           number_format="#,##0.000" if col == 4 else None, bg=bg)
            data_style(ws_m.cell(row=r_idx, column=7,
                                 value=f"=D{r_idx}*{PRICE_CELL}"),
                       number_format="$#,##0.0000", bg=bg)
            ws_m.row_dimensions[r_idx].height = 18

        total_row = first + len(mdf)
        last      = total_row - 1
        for col, v in enumerate([
            "MONTHLY TOTAL", "", f"=SUM(C{first}:C{last})",
            f"=SUM(D{first}:D{last})", "", "",
            f"=SUM(G{first}:G{last})"
        ], 1):
            c = ws_m.cell(row=total_row, column=col, value=v)
            fmt = "$#,##0.0000" if col == 7 else ("#,##0.000" if col == 4 else None)
            data_style(c, bold=True, number_format=fmt, bg=CLR_TOTAL_BG)

        for i, w in enumerate([13, 12, 10, 11, 13, 13, 18], 1):
            ws_m.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
